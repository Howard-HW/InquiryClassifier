import openpyxl
from copy import copy
import pandas as pd
import logging

class ExcelHandler:
    @staticmethod
    def read_excel(file_path, sheet_name):
        try:
            # 엔진을 명시적으로 지정하고 필요한 데이터만 로드
            df = pd.read_excel(
                file_path, 
                sheet_name=sheet_name,
                engine='openpyxl',
                keep_default_na=False  # NA 처리 최적화
            )
            return df
        except Exception as e:
            logging.error(f"Error reading Excel file: {str(e)}")
            raise

    @staticmethod
    def _safe_copy_style(source_cell, target_cell):
        """셀 스타일을 안전하게 복사합니다."""
        try:
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
                
                # 개별 스타일 속성들을 안전하게 복사
                if hasattr(source_cell, 'font'):
                    target_cell.font = copy(source_cell.font)
                if hasattr(source_cell, 'border'):
                    target_cell.border = copy(source_cell.border)
                if hasattr(source_cell, 'fill'):
                    target_cell.fill = copy(source_cell.fill)
                if hasattr(source_cell, 'number_format'):
                    target_cell.number_format = copy(source_cell.number_format)
                if hasattr(source_cell, 'protection'):
                    target_cell.protection = copy(source_cell.protection)
                if hasattr(source_cell, 'alignment'):
                    target_cell.alignment = copy(source_cell.alignment)
        except Exception as e:
            logging.warning(f"Failed to copy cell style: {str(e)}")

    @staticmethod
    def save_excel_with_style(input_file, output_file, sheet_name, data, is_update=False):
        """스타일을 유지하면서 Excel 파일을 저장합니다."""
        try:
            # 원본 파일 로드
            original_wb = openpyxl.load_workbook(input_file)
            original_sheet = original_wb[sheet_name]

            if is_update:
                logging.info(f"Updating original file: {input_file}")
                # 데이터만 업데이트 (2번째 행부터)
                for row_idx, row in enumerate(data.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        original_sheet.cell(row=row_idx, column=col_idx, value=value)
                original_wb.save(input_file)
                logging.info(f"Successfully updated original Excel file: {input_file}")
            else:
                logging.info(f"Creating new file: {output_file}")
                # 새 워크북 생성
                new_wb = openpyxl.Workbook()
                new_sheet = new_wb.active
                new_sheet.title = sheet_name

                try:
                    # 기본 속성 복사 시도
                    new_sheet.sheet_format = copy(original_sheet.sheet_format)
                    new_sheet.sheet_properties = copy(original_sheet.sheet_properties)
                except Exception as e:
                    logging.warning(f"Failed to copy sheet properties: {str(e)}")

                # 열 너비 복사
                try:
                    for column in original_sheet.column_dimensions:
                        new_sheet.column_dimensions[column] = copy(original_sheet.column_dimensions[column])
                except Exception as e:
                    logging.warning(f"Failed to copy column dimensions: {str(e)}")

                # 먼저 헤더(컬럼명) 쓰기
                for col_idx, column_name in enumerate(data.columns, 1):
                    new_cell = new_sheet.cell(row=1, column=col_idx, value=str(column_name))
                    try:
                        original_cell = original_sheet.cell(row=1, column=col_idx)
                        ExcelHandler._safe_copy_style(original_cell, new_cell)
                    except Exception as e:
                        logging.warning(f"Failed to copy header style at column {col_idx}: {str(e)}")

                # 데이터 쓰기 (2번째 행부터)
                for row_idx, row in enumerate(data.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        new_cell = new_sheet.cell(row=row_idx, column=col_idx, value=value)
                        try:
                            original_cell = original_sheet.cell(row=row_idx, column=col_idx)
                            ExcelHandler._safe_copy_style(original_cell, new_cell)
                        except Exception as e:
                            logging.warning(f"Failed to copy style at row {row_idx}, col {col_idx}: {str(e)}")

                try:
                    # 병합된 셀 복사
                    for merged_range in original_sheet.merged_cells:
                        new_sheet.merge_cells(str(merged_range))
                except Exception as e:
                    logging.warning(f"Failed to copy merged cells: {str(e)}")

                try:
                    # 조건부 서식 복사
                    if original_sheet.conditional_formatting:
                        new_sheet.conditional_formatting = copy(original_sheet.conditional_formatting)
                except Exception as e:
                    logging.warning(f"Failed to copy conditional formatting: {str(e)}")

                # 새 파일 저장
                new_wb.save(output_file)
                logging.info(f"Successfully saved new Excel file: {output_file}")

            logging.debug(f"Processed columns: {list(data.columns)}")

        except Exception as e:
            logging.error(f"Error saving Excel file: {str(e)}", exc_info=True)
            raise


    @staticmethod
    def update_excel(file_path, sheet_name, data):
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
            logging.info(f"Successfully updated Excel file: {file_path}, sheet: {sheet_name}")
        except Exception as e:
            logging.error(f"Error updating Excel file: {str(e)}")
            raise